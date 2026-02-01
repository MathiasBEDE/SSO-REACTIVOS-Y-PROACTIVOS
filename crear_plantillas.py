# -*- coding: utf-8 -*-
"""
Script para generar las plantillas Excel de carga masiva
Ejecutar una vez para crear los archivos de plantilla
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Meses en español
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

# Estilos
HEADER_FILL = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def aplicar_estilos_hoja(ws, num_columnas):
    """Aplica estilos a la hoja"""
    # Estilos de header
    for col in range(1, num_columnas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Estilos de datos
    for row in range(2, 14):  # 12 meses
        for col in range(1, num_columnas + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(12, max_length + 2)


def crear_plantilla_reactivos():
    """Crea plantilla para indicadores reactivos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reactivos"
    
    # Headers - solo campos de ENTRADA (no calculados)
    headers = [
        'mes',              # Periodo
        'num_trabajadores', # # Trabajador
        'horas_trabajador', # Horas / Trabajador
        'horas_extras',     # Horas Extras
        'acc_con_baja',     # # Acc. con baja
        'acc_sin_baja',     # # Acc. sin baja
        'enf_ocupacional',  # # Enf. Ocupac
        'dias_perdidos'     # Total Días Perdidos
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Meses con valores por defecto
    for row, mes in enumerate(MESES, 2):
        ws.cell(row=row, column=1, value=mes)
        for col in range(2, len(headers) + 1):
            ws.cell(row=row, column=col, value=0)
    
    aplicar_estilos_hoja(ws, len(headers))
    
    # Guardar
    wb.save('templates/plantilla_reactivos.xlsx')
    print("✓ plantilla_reactivos.xlsx creada")


def crear_plantilla_proactivos():
    """Crea plantilla para indicadores proactivos (7 hojas)"""
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Definir estructura de cada indicador
    indicadores = {
        'IART': ['mes', 'narp', 'nart'],
        'OPAS': ['mes', 'opasp', 'pobp', 'opasr', 'pc'],
        'IDS': ['mes', 'ncsd', 'ncse'],
        'IDPS': ['mes', 'dpsp', 'pp', 'dpsr', 'nas'],
        'IENTS': ['mes', 'nteep', 'nee'],
        'IOSEA': ['mes', 'oseaa', 'oseac'],
        'ICAI': ['mes', 'nmp', 'nmi'],
        'IEF': ['mes', 'capp', 'cape']  # Índice de Eficacia de Formación
    }
    
    for nombre_hoja, columnas in indicadores.items():
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Headers
        for col, header in enumerate(columnas, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Meses y valores por defecto
        for row, mes in enumerate(MESES, 2):
            ws.cell(row=row, column=1, value=mes)
            for col in range(2, len(columnas) + 1):
                ws.cell(row=row, column=col, value=0)
        
        aplicar_estilos_hoja(ws, len(columnas))
    
    # Guardar
    wb.save('templates/plantilla_proactivos.xlsx')
    print("✓ plantilla_proactivos.xlsx creada")


if __name__ == "__main__":
    os.makedirs('templates', exist_ok=True)
    crear_plantilla_reactivos()
    crear_plantilla_proactivos()
    print("\n¡Plantillas creadas exitosamente!")
